from django.shortcuts import render
from django.contrib.staticfiles.storage import staticfiles_storage
import openpyxl

import os
rootPath = os.path.dirname(os.path.dirname(__file__))

def read(request):
    if "GET" == request.method:
        return render(request, 'excelRead/index.html')
    else:
        excel_file = request.FILES["excel_file"]

         # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)


        s_wb = openpyxl.load_workbook(rootPath+'/excelRead/'+staticfiles_storage.url('excelRead/resources/lastChecked.xlsx'))
        s_worksheet = s_wb["Sheet1"]
        s_excel_data = list()
      
        for row in s_worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            s_excel_data.append(row_data)

        result=''
        if excel_data==s_excel_data:
            result='NO change found'
        else:
            result='Change found'

        return render(request, 'excelRead/index.html', {"excel_data":excel_data,'result':result})
